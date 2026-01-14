import os
import sys
import numpy
import pandas
import pathlib
import pubchempy as pcp
import pandas as pds
import PySide6.QtGui
import matplotlib.pyplot as plt
import functions

from PySide6.QtCore import Slot, Signal, QThread
from PySide6.QtWidgets import QPushButton, QWidget, QVBoxLayout, QLabel, \
    QApplication, QMainWindow, QHBoxLayout, QSizePolicy, QFrame, QFileDialog, QGroupBox, QInputDialog, QProgressDialog
from PySide6.QtGui import QPixmap
from PySide6.QtGui import *
from rdkit import Chem
from rdkit.Chem import Draw
from PIL import Image
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from matplotlib.lines import Line2D
import matplotlib.cm as cm

'''Define the UI elements in this part of py file.'''


class MainWindow(QWidget):
    get_file_name = Signal(str)
    get_array = Signal(numpy.ndarray)

    @staticmethod
    def SetPathWay(self):
        self.saving_pathway = QFileDialog.getExistingDirectory(caption='Select the saving directory')[0]
        pass

    @staticmethod
    def buildFunctionButtons(text: str):
        button = QPushButton(text)
        button.setMinimumSize(100, 100)
        button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Expanding)
        return button

    def buildStatus(self):
        self.status_container = QFrame()
        self.status_layout = QHBoxLayout()
        self.status_container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.label_saved_index = QLabel('Saved file: ')

        self.status_layout.addWidget(self.label_file_name)
        self.status_layout.addWidget(self.label_saved_index)
        self.status_layout.addWidget(self.label_AMD)
        self.status_container.setLayout(self.status_layout)

        self.status_container.setFrameStyle(QFrame.StyledPanel | QFrame.Raised)
        return self.status_container

    def buildAnalyzingZone(self):
        self.cmp_structure_label.setFixedSize(200, 200)
        self.cmp_structure_label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        analyze_zone_layout = QHBoxLayout()
        analyze_label_widget = QGroupBox('Compound Information')
        analyze_label_layout = QVBoxLayout()
        analyze_label_layout.addWidget(self.cmp_stay_time)
        analyze_label_layout.addWidget(self.cmp_name_label)
        analyze_label_layout.addWidget(self.cmp_formula_label)
        analyze_label_layout.addWidget(self.cmp_SMILES_label)
        analyze_label_layout.addWidget(self.cmp_carbon_number_label)
        analyze_label_layout.addWidget(self.cmp_Category_label)
        analyze_label_widget.setLayout(analyze_label_layout)

        self.picture_zone = QGroupBox('Structure')
        self.picture_zone.setFixedSize(250, 250)
        self.picture_zone.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        picture_layout = QHBoxLayout()
        picture_layout.addWidget(self.cmp_structure_label)
        self.picture_zone.setLayout(picture_layout)
        analyze_zone_layout.addWidget(analyze_label_widget)
        analyze_zone_layout.addWidget(self.picture_zone)
        return analyze_zone_layout

    def buildHistogramZone(self):
        self.histogram_frame = QGroupBox('The GC-MS Graph')
        this_layout = QVBoxLayout()
        this_layout.addWidget(self.view)
        self.histogram_frame.setLayout(this_layout)
        return

    def buildVisualization(self):
        visualization_layout = QVBoxLayout()
        visualization_container = QFrame()

        status_container = self.buildStatus()
        self.buildHistogramZone()

        visualization_layout.addWidget(status_container)
        visualization_layout.addWidget(self.histogram_frame)
        visualization_layout.addLayout(self.buildAnalyzingZone())
        visualization_container.setLayout(visualization_layout)
        return visualization_container

    def __init__(self):
        super().__init__()
        self.base_pathway = str(pathlib.Path.cwd())
        print(self.base_pathway)
        self.saving_pathway = self.base_pathway + '\\saving file'
        self.temporary_pathway = self.base_pathway + '\\temporary file'
        if not os.path.exists(self.saving_pathway):
            os.mkdir(self.saving_pathway)
        if not os.path.exists(self.temporary_pathway):
            os.mkdir(self.temporary_pathway)
        self.using_paths = [self.temporary_pathway, self.saving_pathway]





        self.view = FigureCanvasQTAgg(Figure(figsize=(5, 3)))
        self.axes = self.view.figure.subplots()
        self.this_layout = None

        self.picture_zone = None
        self.label_saved_index = None
        self.label_file_name = QLabel('Analyzed file:')
        self.label_AMD = QLabel('AMD:')
        self.histogram_frame = None
        self.status_container = None
        self.status_layout = None
        self.task_1 = None
        self.task_1: TaskFindMolecular

        # Analyzing Zone's Widgets, 'cmp' refer to compound.
        self.cmp_stay_time = QLabel('Stay Time:')
        self.cmp_name_label = QLabel('Name:')
        self.cmp_formula_label = QLabel('Formula:')
        self.cmp_SMILES_label = QLabel('SMILES:')
        self.cmp_carbon_number_label = QLabel('Carbon Number:')
        self.cmp_Category_label = QLabel('Category:')
        self.cmp_structure_label = QLabel()

        self.setMinimumSize(800, 600)
        # Set the layout of the main window.
        main_layout = QHBoxLayout()

        # Function buttons' container and layout.
        function_button_container = QWidget()
        function_button_layout = QVBoxLayout()

        # The main functional buttons' initialization.
        self.button_analyse = self.buildFunctionButtons('ANALYZE')
        self.button_analyse.clicked.connect(self.startAnalyze)
        self.button_select_peak = self.buildFunctionButtons('SELECT PEAK')
        self.button_search = self.buildFunctionButtons('SEARCH ')
        self.button_category = self.buildFunctionButtons('CATEGORY')
        self.button_visualize = self.buildFunctionButtons('VISUALIZE')

        self.button_select_peak.clicked.connect(self.selectPeakLogic)
        self.button_search.clicked.connect(self.startSearch)
        self.button_category.clicked.connect(self.startAnalyze)
        self.button_visualize.clicked.connect(self.startAnalyze)

        self.button_analyse.setToolTip('Start the analyze. Need the raw txt data from GC-MS as input.')

        function_button_layout.addWidget(self.button_analyse)
        function_button_layout.addWidget(self.button_select_peak)
        function_button_layout.addWidget(self.button_search)
        function_button_layout.addWidget(self.button_category)
        function_button_layout.addWidget(self.button_visualize)


        self.image_label = QLabel()
        pixmap = QPixmap('C:/Users/ranwu/Documents/Oil Maggot/bin/bocue.jpg')
        self.image_label.setPixmap(pixmap)
        self.image_label.setScaledContents(True)

        main_layout.addLayout(function_button_layout)
        visualization_container = self.buildVisualization()
        main_layout.addWidget(visualization_container)
        main_layout.setStretch(0, 100)
        function_button_layout.setSpacing(0)
        self.setLayout(main_layout)

        self.setWindowTitle('OilMaggot')
        self.setWindowIcon(PySide6.QtGui.QIcon('resources/icons/icon-main.exif'))

    @Slot()
    def startAnalyze(self):
        file_paths = QFileDialog.getOpenFileNames(caption='Select A txt File')[0]
        self.get_file_name.connect(self.changeTheStatus)
        self.get_file_name.emit(file_paths)
        self.task_1 = TaskFindMolecular(file_paths, using_paths=self.using_paths)
        self.task_1.started.connect(self.analyseButtonDisable)
        self.task_1.on_end.connect(self.analyseButtonAble)
        self.task_1.get_the_df.connect(self.initializeHistogram)
        self.task_1.get_the_peak.connect(self.updateTheCompoundInform)
        self.task_1.update_the_histogram.connect(self.UpdateHistogram)
        self.task_1.start()

    @Slot()
    def selectPeakLogic(self):
        # 定义选项列表
        items = ["Highest Possibility", "Custom"]

        # 参数依次为：父窗口、标题、提示语、选项列表、默认索引、是否可编辑
        item, ok = QInputDialog.getItem(self, "Select Peak Logic",
                                        "Please choose the selection logic:",
                                        items, 0, False)

        if ok and item:
            # 暂时只打印结果，不执行后续功能
            print(f"User selected: {item}")

    @Slot()
    def startSearch(self):
        # 1. 选择输入 Excel 文件
        file_info = QFileDialog.getOpenFileName(self, 'Select the Excel file', '', 'Excel files(*.xlsx , *.xls)')
        file_path = file_info[0]
        if not file_path:
            return

        # 获取路径和文件名信息
        self.using_paths = os.path.split(file_path)
        self.filename = os.path.splitext(self.using_paths[1])[0]
        output_filename = self.filename + '_Search_Result.xlsx'
        output_path = os.path.join(self.using_paths[0], output_filename)

        # 2. 读取 Excel 数据
        try:
            input_df = pds.read_excel(file_path)
            # 默认读取第一列作为搜索词
            search_list = input_df.iloc[:, 0].tolist()
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return

        total_count = len(search_list)
        if total_count == 0:
            return

        # 3. 初始化进度条对话框
        progress = QProgressDialog("Searching PubChem...", "Cancel", 0, total_count, self)
        progress.setWindowTitle("Search Progress")
        progress.setWindowModality(Qt.WindowModal)  # 模态对话框，锁定父窗口
        progress.setMinimumDuration(0)  # 立即显示
        progress.show()

        # 4. 初始化 Excel 保存对象 (复用你现有的 openpyxl 逻辑)
        from openpyxl import Workbook
        import re
        wb = Workbook()
        ws = wb.active
        ws.title = "Search Results"
        ws.append(['Query Name', 'IUPAC Name', 'CID', 'CAS', 'SMILES', 'Formula'])

        # 5. 执行搜索循环
        for i, query_item in enumerate(search_list):
            # 更新进度条
            progress.setValue(i)
            progress.setLabelText(f"Searching ({i}/{total_count}): {query_item}")

            # 处理用户点击“取消”的情况
            if progress.wasCanceled():
                print("Search cancelled by user.")
                break

            # 保持界面响应
            QApplication.processEvents()

            query_str = str(query_item).strip()
            if not query_str or query_str == 'nan':
                ws.append([query_str, "Empty query"])
                continue

            try:
                # 核心搜索逻辑（与 ANALYZE 保持一致）
                searching_results = pcp.get_compounds(query_str, 'name')
                if searching_results:
                    comp = searching_results[0]
                    # 获取 CAS (复用原有正则逻辑)
                    cas_id = "N/A"
                    synonyms = pcp.get_synonyms(comp.cid, 'cid')
                    if synonyms:
                        all_syns = " ".join(synonyms[0].get('Synonym', []))
                        cas_list = re.findall(r'\d{2,7}-\d{2}-\d', all_syns)
                        if cas_list:
                            cas_id = cas_list[0]

                    ws.append([
                        query_str,
                        comp.iupac_name if comp.iupac_name else "N/A",
                        comp.cid,
                        cas_id,
                        comp.isomeric_smiles,
                        comp.molecular_formula
                    ])
                else:
                    ws.append([query_str, "Not Found"])
            except Exception as e:
                print(f"Error searching {query_str}: {e}")
                ws.append([query_str, f"Error: {str(e)}"])

        # 6. 完成并保存
        progress.setValue(total_count)
        try:
            wb.save(output_path)
            print(f"Finished! File saved at: {output_path}")
        except Exception as e:
            print(f"Save error: {e}")



    @Slot(pandas.DataFrame, float)
    def initializeHistogram(self, arg, arg1):
        custom_lines = [Line2D([0], [0], color='orange', lw=4),
                        Line2D([0], [0], color='lime', lw=4),
                        Line2D([0], [0], color='purple', lw=4),
                        Line2D([0], [0], color='cyan', lw=4),
                        Line2D([0], [0], color='black', lw=4)]

        self.axes.legend(custom_lines, ['Paraffin', 'Olefin', 'Aromatic', 'Naphthene', 'Hetroatomic'])
        y = arg.iloc[:, 2]
        x = [x for x in range(0, len(y))]
        bar_colors = ['grey']
        self.axes.bar(x, y, color=bar_colors, width=0.8)
        self.axes.set_xlabel('Sequence')
        self.axes.set_ylabel('Proportion')
        self.view.draw()
        self.label_AMD.setText('AMD: ' + str(arg1))
        return

    @Slot(pandas.DataFrame, list)
    def UpdateHistogram(self, arg1, arg2):
        self.axes.clear()
        custom_lines = [Line2D([0], [0], color='orange', lw=4),
                        Line2D([0], [0], color='lime', lw=4),
                        Line2D([0], [0], color='purple', lw=4),
                        Line2D([0], [0], color='cyan', lw=4),
                        Line2D([0], [0], color='black', lw=4)]

        self.axes.legend(custom_lines, ['Paraffin', 'Olefin', 'Aromatic', 'Naphthene', 'Hetroatomic'])
        print('start drawing')
        now_sequence = len(arg1)
        y = arg1.iloc[:, 2]
        x = [x for x in range(0, len(y))]
        bar_colors = arg2
        self.axes.bar(x, y, color=bar_colors, width=0.8)
        self.axes.set_xlabel('Sequence')
        self.axes.set_ylabel('Proportion')
        self.view.draw()
        return

    @Slot()
    def Open(self):

        pass
    @Slot()
    def analyseButtonAble(self):
        self.button_analyse.setEnabled(True)

    @Slot()
    def analyseButtonDisable(self):
        self.button_analyse.setEnabled(False)

    @Slot(str)
    def changeTheStatus(self, arg):
        self.label_file_name.setText('Analyzed file:' + arg)

    @Slot(list)
    def updateTheCompoundInform(self, arg):
        print('In Slot')
        self.cmp_name_label.setText('Name: ' + arg[0])
        self.cmp_formula_label.setText('Formula: ' + arg[1])
        self.cmp_SMILES_label.setText('SMILES: ' + arg[2])
        self.cmp_carbon_number_label.setText('Carbon Number: ' + arg[3])
        self.cmp_Category_label.setText('Category: ' + arg[4])
        self.cmp_stay_time.setText('Stay Time: ' + arg[5])

        mol = Chem.MolFromSmiles(arg[2])
        options = Draw.MolDrawOptions()
        options.addAtomIndices = True
        structure_img = Draw.MolToImage(mol, (250, 300), options=options)
        img = structure_img.toqpixmap()
        self.cmp_structure_label.setPixmap(img)
        self.cmp_structure_label.setScaledContents(True)


class TaskFindMolecular(QThread):
    on_end = Signal()
    get_the_peak = Signal(list)
    get_the_df = Signal(pandas.DataFrame, int)
    get_the_result = Signal(pandas.Series)
    update_the_histogram = Signal(pandas.DataFrame, list)

    def __init__(self, file_paths, using_paths):
        super().__init__()
        self.file_paths = file_paths
        self.using_paths = using_paths
        self.get_the_result.connect(self.OutputResult)
        self.temporary_path = self.using_paths[0] + '\\test excel.xlsx'
        self.filename = self.GetFileName(self, file_paths[0])

    def run(self):
        for file_path in self.file_paths:
            functions.TxtMachine.delete_useless(file_path, self.using_paths[0])
            txt_array = functions.TxtMachine.txt_to_numpy_array(self.using_paths[0] + 'output_file.txt')
            peaks = functions.TxtMachine.get_peak_collection(txt_array)
            functions.PeakMachine.peaksToExcel(self.temporary_path, peaks)
            df = pds.read_excel(self.temporary_path)
            AMD = self.CalculateAMD(self, df)
            self.get_the_df.emit(df, AMD)
            self.findMolecular(self, df)

        self.on_end.emit()

    @staticmethod
    def DecideColor(self, category: str):
        color = ''
        if category == 'Heteroatomic':
            color = 'black'
        elif category == 'Paraffin':
            color = 'orange'
        elif category == 'Aromatic':
            color = 'purple'
        elif category == 'Olefin':
            color = 'lime'
        elif category == 'Naphthene':
            color = 'cyan'
        elif category == 'Unkonwn':
            color = 'grey'
        return color

    @staticmethod
    def CalculateAMD(self, df: pandas.DataFrame):
        AMD = 0
        total_value = 0
        for i in range(len(df)):
            total_value += (df.iloc[i, 2] * df.iloc[i, 5]) / 100
        AMD = total_value
        return AMD

    @staticmethod
    def DrawResult(self, df: pandas.DataFrame):

        return

    @staticmethod
    def findMolecular(self, df):
        # 加载Excel文件
        histogram_color_list = ['grey' for i in range(len(df))]
        df = pds.read_excel(self.temporary_path)
        df.insert(loc=7, column='Molecular Formula', value='Not Found', allow_duplicates=False)
        df.insert(loc=8, column='SMILES', value='Not Found', allow_duplicates=False)
        df.insert(loc=9, column='Carbon Number', value='Not Found', allow_duplicates=False)
        df.insert(loc=10, column='CateGory', value='Not Found', allow_duplicates=False)
        df.insert(loc=11, column='HIC', value='Not Found', allow_duplicates=False)
        df.insert(loc=12, column='LIC', value='Not Found', allow_duplicates=False)
        df.insert(loc=13, column='HC', value='Not Found', allow_duplicates=False)
        df.insert(loc=14, column='LC', value='Not Found', allow_duplicates=False)

        sequence = 0
        last_formula = 'C1H4'
        last_smiles = 'C'
        name = 'Default'
        saving_path = self.using_paths[1] + '\\' + self.filename + '.xlsx'
        for i in range(0, len(df)):
            print('\nstart searching MOLECULAR STRUCTURE...: ' + df.iat[i, 3])
            try:
                name: str
                name = df.iat[i, 3]
                stay_time = str(df.iat[i, 1])
                name = name.rstrip()
                compound = pcp.get_compounds(identifier=df.iat[i, 3], namespace='name')[0]
                print(compound.to_dict().get('record').get('props'))
                print('Yes')
                smile = [i for i in compound.to_dict().get('record').get('props') if i['urn']['label'] == 'SMILES']
                smile = smile[0]['value']['sval']
                print(smile)
                category_info = functions.MolecularCoping.getCategory(compound.molecular_formula,
                                                                     smile,
                                                                     name)
                category = str(category_info[0])
                print('category: ' + str(category))
                ui_list = [name, compound.molecular_formula, smile,
                           functions.MolecularCoping.getCarbonNumber(compound.molecular_formula)[0],
                           category, stay_time]

                # Emit a self_defined signal when get the result from PyPubchem.
                now_color = self.DecideColor(self, category)
                histogram_color_list[sequence] = now_color
                self.get_the_peak.emit(ui_list)
                self.update_the_histogram.emit(df, histogram_color_list)
                sequence += 1
                df.iat[i, 7] = compound.molecular_formula
                df.iat[i, 8] = compound.canonical_smiles
                df.iat[i, 9] = functions.MolecularCoping.getCarbonNumber(compound.molecular_formula)[0]
                df.iat[i, 10] = category_info[0]
                df.iat[i, 11] = category_info[2]
                df.iat[i, 12] = category_info[3]
                df.iat[i, 13] = category_info[4]
                df.iat[i, 14] = category_info[5]

                last_formula = compound.molecular_formula
                last_smiles = compound.canonical_smiles

                df.to_excel(saving_path, sheet_name='End', index=False)

            except:

                print('search err, return not found')
                df.iat[i, 7] = last_formula
                df.iat[i, 8] = last_smiles
                df.iat[i, 9] = functions.MolecularCoping.getCarbonNumber(last_formula)[0]
                # 分类部分
                df.iat[i, 10] = category_info[0]
                df.iat[i, 11] = category_info[2]
                df.iat[i, 12] = category_info[3]
                df.iat[i, 13] = category_info[4]
                df.iat[i, 14] = category_info[5]
                sequence += 1

                df.to_excel(saving_path, sheet_name='End', index=False)

        self.get_the_result.emit(df)
        return

    @Slot(pandas.DataFrame)
    def OutputResult(self, arg1):
        filename = self.file_paths[0]
        import matplotlib.pyplot as mtplt
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        from matplotlib.ticker import MaxNLocator

        result_pivot = pds.pivot_table(arg1, index=['Carbon Number'], columns=['CateGory'], values='N%', aggfunc='sum',
                                       margins=True)
        result_pivot.to_excel(self.using_paths[1] + '\\' + self.filename + '(pivot).xlsx', sheet_name='End')

        wb = load_workbook(self.using_paths[1] + '\\' + self.filename + '(pivot).xlsx')
        ws = wb.worksheets[0]
        number_distribution = arg1.groupby('Carbon Number')['N%'].sum()
        number_distribution.index = number_distribution.index.astype(int)
        number_distribution.sort_values()
        category_distribution = arg1.groupby('CateGory')['N%'].sum()
        x1 = number_distribution.index
        x2 = category_distribution.index
        y1 = number_distribution.values
        y2 = category_distribution.values

        print(x1)
        print(y1)
        mtplt.figure(figsize=(5, 4))
        mtplt.bar(x1, y1, color='green')
        mtplt.gca().xaxis.set_major_locator(MaxNLocator(integer=True))
        mtplt.xlabel('Carbon Number')
        mtplt.ylabel('N%')
        mtplt.title('Carbon Number - N%')
        cur_path = 'bar_chart1.png'
        mtplt.savefig(cur_path, dpi=800)
        mtplt.xticks(range(min(x1) - 1, max(x1) + 1))
        mtplt.close()
        img = Image(cur_path)
        img.width = 450
        img.height = 300
        ws.add_image(img, 'G1')

        mtplt.figure(figsize=(5, 4))
        mtplt.bar(x2, y2, color='orange')
        mtplt.xlabel('Category')
        mtplt.ylabel('N%')
        mtplt.title('Category - N%')
        cur_path = 'bar_chart2.png'
        mtplt.savefig(cur_path, dpi=800)
        mtplt.close()
        img = Image(cur_path)
        img.width = 450
        img.height = 300
        ws.add_image(img, 'G20')

        print(number_distribution)
        print(category_distribution)

        wb.save(self.using_paths[1] + '\\' + self.filename + '(pivot).xlsx')

        return

    @staticmethod
    def GetFileName(self, filepath: str):
        filename = ''
        position = 0
        last_slash = 0
        for i in filepath:
            if i == '/':
                last_slash = position
            position += 1
        filename = filepath[last_slash + 1:len(filepath) - 4]
        return filename




if __name__ == '__main__':
    app = QApplication()
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec())
